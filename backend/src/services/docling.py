from fastapi import UploadFile, HTTPException, status
from loguru import logger
import openpyxl
from pypdf import PdfReader
import io

async def process_document_with_docling(file: UploadFile) -> str:
    """
    Simulates a "Docling" service.
    Processes an uploaded file (PDF, XLSX, TXT, MD) and extracts text content.
    """
    content_type = file.content_type
    logger.info(f"Processing file '{file.filename}' with content type: {content_type}")
    
    file_bytes = await file.read()
    text_content = ""

    try:
        if content_type == "application/pdf":
            reader = PdfReader(io.BytesIO(file_bytes))
            for page in reader.pages:
                text_content += page.extract_text() or ""
        
        elif content_type in ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/vnd.ms-excel"]:
            workbook = openpyxl.load_workbook(io.BytesIO(file_bytes))
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_content += f"\n--- Sheet: {sheet_name} ---\n"
                for row in sheet.iter_rows():
                    row_text = "\t".join([str(cell.value) if cell.value is not None else "" for cell in row])
                    text_content += row_text + "\n"

        elif content_type in ["text/plain", "text/markdown"]:
            text_content = file_bytes.decode("utf-8")

        else:
            raise HTTPException(
                status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
                detail=f"Unsupported file type: {content_type}. Please upload PDF, Excel, TXT, or MD files."
            )
            
        logger.success(f"Successfully extracted text from '{file.filename}'.")
        return text_content.strip()

    except Exception as e:
        logger.error(f"Failed to process file '{file.filename}': {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to process file: {str(e)}"
        )